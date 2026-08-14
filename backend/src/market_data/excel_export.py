"""Excel export for OHLCV data (tasks 5.1, 5.2)."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from market_data.models import OHLCV_COLUMNS, Series

logger = logging.getLogger(__name__)

_HEADER = [*OHLCV_COLUMNS, "open_time_iso"]


def _with_iso(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    # Excel cannot store tz-aware datetimes; keep tz-naive UTC.
    out["open_time_iso"] = pd.to_datetime(out["open_time"], unit="ms", utc=True).dt.tz_localize(None)
    return out[_HEADER]


# -- 5.1: one-shot export --------------------------------------------------
def export_frame(frame: pd.DataFrame, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    _with_iso(frame).to_excel(path, index=False)
    logger.info("Exported %d rows to %s.", len(frame), path)
    return path


def export_series(store, series: Series, out_dir: Path) -> list[Path]:  # noqa: ANN001
    """Export a series to one .xlsx per UTC calendar day."""
    frame = store.read(series)
    if frame.empty:
        return []
    day = pd.to_datetime(frame["open_time"], unit="ms", utc=True).dt.strftime("%Y-%m-%d")
    base = out_dir / series.relative_path()
    paths: list[Path] = []
    for d, group in frame.assign(_day=day).groupby("_day"):
        path = base / f"{d}.xlsx"
        export_frame(group.drop(columns="_day"), path)
        paths.append(path)
    return paths


# -- 5.2: real-time batched appender --------------------------------------
class ExcelAppender:
    """Append candles to an .xlsx in batches to avoid per-row write overhead."""

    def __init__(self, path: Path, *, batch_size: int = 500) -> None:
        self._path = Path(path)
        self._batch_size = batch_size
        self._buffer: list[list] = []
        self._ws = None
        self._wb: Workbook | None = None

    def __enter__(self) -> "ExcelAppender":
        self._path.parent.mkdir(parents=True, exist_ok=True)
        if self._path.exists():
            self._wb = load_workbook(self._path)
            self._ws = self._wb.active
        else:
            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.append(_HEADER)
        return self

    def append(self, frame: pd.DataFrame) -> None:
        if frame.empty:
            return
        for row in _with_iso(frame).itertuples(index=False):
            self._buffer.append(list(row))
        if len(self._buffer) >= self._batch_size:
            self.flush()

    def flush(self) -> None:
        if not self._buffer or self._ws is None or self._wb is None:
            return
        for row in self._buffer:
            self._ws.append(row)
        self._wb.save(self._path)
        logger.info("Flushed %d rows to %s.", len(self._buffer), self._path)
        self._buffer.clear()

    def __exit__(self, *_exc: object) -> None:
        self.flush()
        if self._wb is not None:
            self._wb.close()
